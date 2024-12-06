from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from typing import List, Dict, Optional

class MyWorkbook:
    def __init__(self, wb_file: str, first_page: str = '') -> None:
        """
        Initialize the workbook and set the first sheet title.
        """
        self.wb = Workbook()
        self.wb_file = wb_file
        self.ws = self.wb.active
        self.ws.title = first_page if first_page else 'First Page'
    
    def create_sheet(self, ws_title: str) -> None:
        """
        Create a new sheet with the given title.
        """
        self.wb.create_sheet(ws_title)
    
    def change_active_sheet(self, ws_title: str) -> None:
        """
        Change the active sheet to the one with the given title.
        """
        self.ws = self.wb[ws_title]
        
    def add_header(self, types: List[str], title: str) -> None:
        """
        Add a header to the active sheet.
        """
        self.ws.append([title])
        self.ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=len(types))
        self.ws.append([''])
        self.ws.append(types)
    
    def format_header(self) -> None:
        """
        Format the header row with center alignment and bold font.
        """
        for col in range(1, self.ws.max_column + 1):
            for row in range(1, self.ws.max_row + 1):
                cell = self.ws.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal='center')
                cell.font = Font(bold=True)
                
    def set_border(self, side: Optional[Side] = None, start: int = 1) -> None:
        """
        Set borders for the cells in the active sheet.
        """
        side = side or Side(border_style='thin', color='8A0101')
        max_col = self.ws.max_column
        max_row = self.ws.max_row

        # Main columns
        for col in range(1, max_col + 1):
            self.ws.cell(row=start, column=col).border = Border(top=side, bottom=side, left=side, right=side)
        
        # Other cells
        for col in range(1, max_col + 1):
            for row in range(start + 1, max_row):
                self.ws.cell(row=row, column=col).border = Border(left=side, right=side)
        
        # Bottom row
        for col in range(1, max_col + 1):
            self.ws.cell(row=max_row, column=col).border = Border(left=side, bottom=side, right=side)
    
    def add_main_data(self, data: List[Dict[str, str]]) -> None:
        """
        Add main data to the active sheet.
        """
        for record in data:
            self.ws.append(list(record.values()))
        
    def edit_width_for_columns(self, list_of_excel_column: List[str], min_row: int) -> None:
        """
        Adjust the width of columns based on the content.
        """
        for letter in list_of_excel_column:
            max_width = 0
            for row_number in range(min_row, self.ws.max_row + 1):
                value = self.ws[f'{letter}{row_number}'].value
                if value and len(value) > max_width:
                    max_width = len(value)
            self.ws.column_dimensions[letter].width = max_width + 1
        self.edit_height_for_row(min_row)
            
    def edit_height_for_row(self, min_row: int = 1) -> None:
        """
        Adjust the height of rows based on the content.
        """
        for row in range(min_row, self.ws.max_row + 1):
            max_height = 0
            for col in range(1, self.ws.max_column + 1):
                value = self.ws.cell(row=row, column=col).value
                if value and isinstance(value, str) and len(value.splitlines()) > 1:
                    max_height = max(max_height, len(value.split('\n')))
            self.ws.row_dimensions[row].height = max_height if max_height > 0 else None
        
    def save_workbook(self) -> None:
        """
        Save the workbook to the specified file.
        """
        self.wb.save(self.wb_file)