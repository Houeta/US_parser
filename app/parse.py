import csv
import os
import requests
import time
from argparse import ArgumentParser, Namespace

from bs4 import BeautifulSoup
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from pprint import pprint
import pandas as pd
import fake_useragent
import urllib

from my_parser import UsersideParser, AbillsParser

parser = ArgumentParser()
# group = parser.add_mutually_exclusive_group()

parser.add_argument('-s', '--start', type=str, help="Start timestamp")
parser.add_argument('-e', '--end', type=str, help="End timestamp")
# group.add_argument('-r', '--verbose', action='count', help='Produce verbose output. Use -vv for extra verbose output.', default=0)
args: Namespace = parser.parse_args()

# load_dotenv('.env')
us_auth_link = 'http://us3.radionet.com.ua/oper/index.php?w=555'
abills_auth_link = 'https://bill-admin2.radionet.com.ua:9443/admin/index.cgi?DOMAIN_ID=&language=english'

userside_payload = {
    "action": "login",
    "username": os.environ.get("user"),
    "password": os.environ.get("pass"),
}

us_petrivskyy_params = {
    "core_section": 'task_list',
    'filter_selector0': 'task_state',
    'task_state0_value': 2,
    'filter_selector1': 'task_staff',
    'employee_find_input': None,
    'employee_id1': 63,
    'filter_selector2': 'date_finish',
    'date_finish2_value2': 1,
    'date_finish2_date1': args.start,
    'date_finish2_date2': args.end,
}

abills_payload = {
    "DOMAIN_ID": "",
    "REFERER": "https://bill-admin2.radionet.com.ua:9443/admin/index.cgi?DOMAIN_ID=&language=english",
    "LOGIN": "1",
    "language": "english",
    "user": os.environ.get("user"),
    "passwd": os.environ.get("pass"),
    "logined": "Enter",
}

def save_file(items, path):
    with open(path, 'w', encoding='utf8', newline='') as f:
        writer = csv.writer(f, delimiter=',')
        writer.writerow(['Історія підключень за 11-2023p. (Петрівський Володимир Романович)'])
        writer.writerow(['Дата','Тип задачі', '№ договору', 'ПІБ', 'Адреса'])
        for item in items:
            writer.writerow([item['exec_date'], item['task_type'], item['contract_id'], item['name'], item['address']])

def save_in_excel_file(data, title, excel_workfile):
    wb = Workbook()
    ws = wb.active
    ws.title = title
    
    headings = list(data[0].keys())
    
    ws.append([f"Історія підключень за {us_petrivskyy_params['date_finish2_date1']}-{us_petrivskyy_params['date_finish2_date2']} (Петрівський Володимир Романович)", ])
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=len(headings))
    ws.append([""])
    ws.append(headings)
    
    for counter in range(len(data)):
        metadata = list(data[counter].values())
        ws.append(metadata)
    
    for letter in ['A', 'B', 'C', 'D', 'E']:
        max_width = 0
        
        for row_number in range(2, ws.max_row + 1):
            if not ws[f'{letter}{row_number}'].value:
                continue    
            elif len(ws[f'{letter}{row_number}'].value) > max_width:
                max_width = len(ws[f'{letter}{row_number}'].value)
                
        ws.column_dimensions[letter].width = max_width + 1
    
    wb.save(excel_workfile)
    
def edit_width_for_columns(excel_workfile, ws):
    for letter in ['A', 'B', 'C', 'D', 'E']:
        max_width = 0
        
        for row_number in range(1, ws.max_row + 1):
            if len(ws[f'{letter}{row_number}'].value) > max_width:
                max_width = len(ws[f'{letter}{row_number}'].value)
                
        ws.column_dimensions[letter].width = max_width + 1

if __name__ == '__main__':
    us_parser = UsersideParser(us_auth_link, userside_payload, us_petrivskyy_params)
    abills_parser = AbillsParser(abills_auth_link, abills_payload, params=None)
    us_dicts = us_parser.parse()
    us_dicts.reverse()
    excel_workfile = f"excel_data/tasks-{us_petrivskyy_params['date_finish2_date1']}-{us_petrivskyy_params['date_finish2_date2']}.xlsx"
    
    only_new_conn = []
    for us_dict in us_dicts:
        if 'Новое подключение' in us_dict['Task type']:
            us_dict["Contract ID"] = abills_parser.get_contract_id(us_dict["Firstname, Lastname"])
            # pprint(us_dict)
            only_new_conn.append(us_dict)
            
    save_in_excel_file(only_new_conn, 'New Connections', excel_workfile)
    #edit_width_for_columns(excel_workfile)

            
    #print(session.get(url=URLs[0], data=petrivskyy_payload, headers=header).text)